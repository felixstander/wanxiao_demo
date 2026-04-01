#!/usr/bin/env python3
"""
CSV to Prompt Generator
读取CSV文件，将字段名和内容拼接为提示词
只使用Python基础库
"""

import argparse
import csv
import sys
from dataclasses import dataclass, field, fields
from pathlib import Path


def _check_openpyxl():
    """检查 openpyxl 是否已安装"""
    try:
        import openpyxl

        return True
    except ImportError:
        return False


@dataclass
class BusinessProvidedFields:
    """产品/业务提供的字段定义"""

    description: str = field(default="", metadata={"description": "使用场景描述"})
    steps: str = field(default="", metadata={"description": "SOP流程(业务/产品提供)"})
    judege_logic: str = field(
        default="", metadata={"description": "步骤分流逻辑(业务/产品提供)"}
    )
    is_inverse: str = field(
        default="", metadata={"description": "需用户反馈内容(业务/产品提供)"}
    )
    inverse_content: str = field(
        default="", metadata={"description": "反馈内容缺失是否反问(业务/产品提供)"}
    )
    related_question: str = field(
        default="", metadata={"description": "建议关联问(业务/产品提供)"}
    )


@dataclass
class DevProvidedFields:
    """开发提供的字段定义"""

    tool_method: str = field(
        default="", metadata={"description": "接口工具（开发提供）"}
    )
    tool_input: str = field(
        default="", metadata={"description": "接口入参（开发提供）"}
    )
    tool_output: str = field(
        default="", metadata={"description": "接口出参（开发提供）"}
    )


@dataclass
class SkillDataRow:
    """完整的 Skill 数据行，合并业务和开发字段"""

    business: BusinessProvidedFields = field(default_factory=BusinessProvidedFields)
    dev: DevProvidedFields = field(default_factory=DevProvidedFields)


def _get_allowed_field_names():
    """获取所有允许的字段名（来自两个 dataclass）"""
    business_fields = [f.name for f in fields(BusinessProvidedFields)]
    dev_fields = [f.name for f in fields(DevProvidedFields)]
    return set(business_fields + dev_fields)


def _parse_row_to_dataclass(headers, row_data):
    """将单行数据解析为 SkillDataRow"""
    allowed_fields = _get_allowed_field_names()

    business_data = {}
    dev_data = {}

    business_field_names = {f.name for f in fields(BusinessProvidedFields)}
    dev_field_names = {f.name for f in fields(DevProvidedFields)}

    for i, header in enumerate(headers):
        if header not in allowed_fields:
            continue
        value = row_data[i] if i < len(row_data) else ""

        if header in business_field_names:
            business_data[header] = value
        elif header in dev_field_names:
            dev_data[header] = value

    return SkillDataRow(
        business=BusinessProvidedFields(**business_data),
        dev=DevProvidedFields(**dev_data),
    )


def _read_csv_single_sheet(file_path):
    """读取单个CSV文件，返回 (sheet_name, headers, skill_data_rows)"""
    rows = []
    with open(file_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append(row)

    if len(rows) == 0:
        return "Sheet1", [], []

    headers = rows[0]
    data_rows = rows[1:]

    allowed_fields = _get_allowed_field_names()
    filtered_headers = [h for h in headers if h in allowed_fields]

    skill_data_rows = []
    for row in data_rows:
        skill_row = _parse_row_to_dataclass(headers, row)
        skill_data_rows.append(skill_row)

    return "Sheet1", filtered_headers, skill_data_rows


def _read_excel_all_sheets(file_path):
    """读取Excel文件所有sheet，返回 [(sheet_name, headers, skill_data_rows), ...]"""
    if not _check_openpyxl():
        print(
            "错误：读取Excel文件需要安装 openpyxl，请运行: pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)

    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"读取Excel文件时出错：{e}", file=sys.stderr)
        sys.exit(1)

    all_sheets_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])

        if len(rows) == 0:
            continue

        headers = rows[0]
        data_rows = rows[1:]

        allowed_fields = _get_allowed_field_names()
        filtered_headers = [h for h in headers if h in allowed_fields]

        if not filtered_headers:
            continue

        skill_data_rows = []
        for row in data_rows:
            if not any(row):
                continue
            skill_row = _parse_row_to_dataclass(headers, row)
            skill_data_rows.append(skill_row)

        if skill_data_rows:
            all_sheets_data.append((sheet_name, filtered_headers, skill_data_rows))

    wb.close()
    return all_sheets_data


def read_data_file(file_path):
    """读取数据文件（CSV或Excel），返回 [(sheet_name, headers, skill_data_rows), ...]"""
    path = Path(file_path)

    if not path.exists():
        print(f"错误：文件 '{file_path}' 不存在", file=sys.stderr)
        sys.exit(1)

    suffix = path.suffix.lower()

    if suffix in [".xlsx", ".xlsm"]:
        return _read_excel_all_sheets(file_path)
    elif suffix == ".csv":
        sheet_data = _read_csv_single_sheet(file_path)
        return [sheet_data]
    else:
        print(
            f"错误：不支持的文件格式 '{suffix}'，请使用 .csv 或 .xlsx", file=sys.stderr
        )
        sys.exit(1)


def generate_prompt(headers, skill_row):
    """将字段名和 SkillDataRow 数据拼接为提示词，测试集字段放最后"""
    normal_parts = []

    business_fields = {f.name: f for f in fields(BusinessProvidedFields)}
    dev_fields = {f.name: f for f in fields(DevProvidedFields)}

    def get_field_value(header):
        if header in business_fields:
            return getattr(skill_row.business, header, "")
        elif header in dev_fields:
            return getattr(skill_row.dev, header, "")
        return ""

    for header in headers:
        value = get_field_value(header)
        part = f"【{header}】\n{value}"
        normal_parts.append(part)

    return "\n\n".join(normal_parts)


def generate_all_prompts(file_path):
    """生成所有数据行的提示词（按 sheet 分组显示）"""
    all_sheets_data = read_data_file(file_path)

    all_results = []

    for sheet_idx, (sheet_name, headers, skill_data_rows) in enumerate(
        all_sheets_data, 1
    ):
        sheet_sections = []
        sheet_sections.append("=" * 60)
        if "sheet" in sheet_name.lower():
            sheet_sections.append(sheet_name)
        else:
            sheet_sections.append(f"Sheet {sheet_idx}: {sheet_name}")
        sheet_sections.append("=" * 60)

        records = []
        for i, skill_row in enumerate(skill_data_rows):
            prompt = generate_prompt(headers, skill_row)
            records.append(f"--- 第 {i + 1} 条记录 ---\n{prompt}")

        sheet_sections.append("")
        sheet_sections.append("\n\n".join(records))
        all_results.append("\n".join(sheet_sections))

    result = "\n\n".join(all_results)
    return result


def main():
    parser = argparse.ArgumentParser(
        description="读取CSV/Excel文件，将字段名和内容拼接为提示词",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
        示例：
        python3 csv_to_prompt.py /workspace/uploaded/data.csv
        python3 csv_to_prompt.py /workspace/uploaded/data.xlsx
        """,
    )

    parser.add_argument(
        "file", help="数据文件的完整路径（包含 .csv 或 .xlsx 后缀）"
    )

    args = parser.parse_args()

    file_path = Path(args.file)

    if not file_path.exists():
        print(f"错误：文件 '{file_path}' 不存在", file=sys.stderr)
        sys.exit(1)

    skill_input = generate_all_prompts(str(file_path))

    return skill_input


if __name__ == "__main__":
    skill_input = main()
    print(skill_input)
